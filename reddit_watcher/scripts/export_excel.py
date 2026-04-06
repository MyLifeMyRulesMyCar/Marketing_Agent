"""
scripts/export_excel.py — Export Reddit posts to a formatted Excel workbook.

Sheets:
  1. All Posts       — every post across all subreddits
  2. Per-subreddit   — one sheet per subreddit
  3. Summary         — stats per subreddit (post count, avg score, top posts)
"""

import os
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


# ── Color palette ──────────────────────────────────────────────
COLOR_HEADER_BG  = "FF4500"   # Reddit orange
COLOR_HEADER_FG  = "FFFFFF"
COLOR_ALT_ROW    = "FFF5F0"   # light orange tint
COLOR_SUB_HEADER = "FFD8C8"
COLOR_GOOD       = "C6EFCE"   # green highlight (high score)
COLOR_NEUTRAL    = "FFEB9C"   # yellow (medium)
FONT_NAME        = "Arial"


def _header_style():
    return Font(name=FONT_NAME, bold=True, color=COLOR_HEADER_FG, size=11)

def _header_fill():
    return PatternFill("solid", fgColor=COLOR_HEADER_BG)

def _alt_fill():
    return PatternFill("solid", fgColor=COLOR_ALT_ROW)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="top", wrap_text=True)

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_header_row(ws, headers: list[str], col_widths: list[int]):
    ws.append(headers)
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font  = _header_style()
        cell.fill  = _header_fill()
        cell.alignment = _center()
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _post_to_row(post: dict) -> list:
    comments_text = " | ".join(
        f"[{c['score']}] {c['body'][:100]}"
        for c in post.get("top_comments", [])
    )
    return [
        post.get("subreddit", ""),
        post.get("title", ""),
        post.get("score", 0),
        post.get("upvote_ratio", 0),
        post.get("num_comments", 0),
        post.get("author", ""),
        post.get("flair", ""),
        post.get("created_utc", ""),
        post.get("permalink", ""),
        post.get("selftext", ""),
        comments_text,
    ]


HEADERS = [
    "Subreddit", "Title", "Score", "Upvote %", "Comments",
    "Author", "Flair", "Posted At", "Link", "Post Text", "Top Comments"
]
COL_WIDTHS = [15, 55, 10, 12, 12, 18, 15, 18, 50, 40, 60]


def _apply_row_style(ws, row_num: int, score: int, is_alt: bool):
    fill = _alt_fill() if is_alt else PatternFill("solid", fgColor="FFFFFF")
    if score >= 100:
        fill = PatternFill("solid", fgColor=COLOR_GOOD)
    elif score >= 20:
        fill = PatternFill("solid", fgColor=COLOR_NEUTRAL)

    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill      = fill
        cell.font      = Font(name=FONT_NAME, size=10)
        cell.border    = _thin_border()
        cell.alignment = _left() if col in [2, 10, 11] else _center()


def write_all_posts_sheet(wb: Workbook, posts: list[dict]):
    ws = wb.active
    ws.title = "All Posts"
    _write_header_row(ws, HEADERS, COL_WIDTHS)
    ws.row_dimensions[1].height = 30

    for i, post in enumerate(sorted(posts, key=lambda p: p.get("score", 0), reverse=True)):
        row_data = _post_to_row(post)
        ws.append(row_data)
        row_num = i + 2
        _apply_row_style(ws, row_num, post.get("score", 0), i % 2 == 0)
        ws.row_dimensions[row_num].height = 60

        # Make permalink a hyperlink
        link_cell = ws.cell(row=row_num, column=9)
        link_cell.hyperlink = post.get("permalink", "")
        link_cell.font = Font(name=FONT_NAME, size=10, color="0563C1", underline="single")

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"


def write_subreddit_sheets(wb: Workbook, posts: list[dict]):
    by_sub = defaultdict(list)
    for post in posts:
        by_sub[post["subreddit"]].append(post)

    for sub_name, sub_posts in sorted(by_sub.items()):
        ws = wb.create_sheet(title=sub_name[:31])
        _write_header_row(ws, HEADERS, COL_WIDTHS)

        for i, post in enumerate(sorted(sub_posts, key=lambda p: p.get("score", 0), reverse=True)):
            row_data = _post_to_row(post)
            ws.append(row_data)
            row_num = i + 2
            _apply_row_style(ws, row_num, post.get("score", 0), i % 2 == 0)
            ws.row_dimensions[row_num].height = 55

            link_cell = ws.cell(row=row_num, column=9)
            link_cell.hyperlink = post.get("permalink", "")
            link_cell.font = Font(name=FONT_NAME, size=10, color="0563C1", underline="single")

        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"


def write_summary_sheet(wb: Workbook, posts: list[dict]):
    ws = wb.create_sheet(title="Summary")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 55

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"Reddit Trends Summary — fetched {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title_cell.font  = Font(name=FONT_NAME, bold=True, size=14, color=COLOR_HEADER_FG)
    title_cell.fill  = _header_fill()
    title_cell.alignment = _center()
    ws.row_dimensions[1].height = 35

    # Sub-headers
    summary_headers = ["Subreddit", "Posts", "Avg Score", "Avg Comments", "Top Score", "Top Post Title"]
    ws.append(summary_headers)
    for col in range(1, 7):
        cell = ws.cell(row=2, column=col)
        cell.font  = Font(name=FONT_NAME, bold=True, size=11)
        cell.fill  = PatternFill("solid", fgColor=COLOR_SUB_HEADER)
        cell.alignment = _center()
        cell.border = _thin_border()
    ws.row_dimensions[2].height = 25

    by_sub = defaultdict(list)
    for post in posts:
        by_sub[post["subreddit"]].append(post)

    for i, (sub_name, sub_posts) in enumerate(sorted(by_sub.items()), start=3):
        scores   = [p.get("score", 0) for p in sub_posts]
        comments = [p.get("num_comments", 0) for p in sub_posts]
        top_post = max(sub_posts, key=lambda p: p.get("score", 0))

        row = [
            f"r/{sub_name}",
            len(sub_posts),
            round(sum(scores) / len(scores), 1) if scores else 0,
            round(sum(comments) / len(comments), 1) if comments else 0,
            max(scores) if scores else 0,
            top_post.get("title", ""),
        ]
        ws.append(row)
        fill = _alt_fill() if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col in range(1, 7):
            cell = ws.cell(row=i, column=col)
            cell.fill      = fill
            cell.font      = Font(name=FONT_NAME, size=10)
            cell.border    = _thin_border()
            cell.alignment = _center() if col < 6 else _left()
        ws.row_dimensions[i].height = 22

    ws.freeze_panes = "A3"


def export_to_excel(posts: list[dict], output_path: str):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = Workbook()

    print(f"\n  📊 Building Excel workbook...")
    write_all_posts_sheet(wb, posts)
    write_subreddit_sheets(wb, posts)
    write_summary_sheet(wb, posts)

    # Move Summary to front (after All Posts)
    wb.move_sheet("Summary", offset=-(len(wb.sheetnames) - 2))

    wb.save(output_path)
    print(f"  ✅ Saved → {output_path}")
    return output_path